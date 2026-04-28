from flask import Blueprint, render_template, redirect, url_for, flash, current_app, send_file
from app.controllers.alertas import AlertaController

from flask_login import login_required

alertas_bp = Blueprint('alertas', __name__)

@alertas_bp.before_request
@login_required
def require_login():
    pass


@alertas_bp.route('/alertas')
def listar():
    productos_bajos = AlertaController.obtener_productos_bajos()
    ultima_global = AlertaController.obtener_ultima_alerta_global()
    return render_template('alertas/index.html', 
                         productos=productos_bajos, 
                         ultima_global=ultima_global)


@alertas_bp.route('/alertas/enviar', methods=['POST'])
def enviar():
    from app.services.alertas import verificar_stock_y_notificar
    from app.controllers.alertas import AlertaController
    
    app = current_app._get_current_object()
    productos_bajos = AlertaController.obtener_productos_bajos()
    
    try:
        verificar_stock_y_notificar(app)
        AlertaController.guardar_fecha_global()
        AlertaController.guardar_fechas_productos(productos_bajos)
        flash('Alerta enviada y fechas actualizadas', 'success')
    except Exception as e:
        flash(f'Error al enviar alerta: {str(e)}', 'danger')
        current_app.logger.error(f"Error al enviar alerta: {str(e)}")
    
    return redirect(url_for('alertas.listar'))


@alertas_bp.route('/alertas/exportar-excel')
def exportar_excel_alertas():
    """Genera un archivo Excel con los productos que tienen stock bajo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from datetime import datetime

    productos = AlertaController.obtener_productos_bajos()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Alertas Stock Bajo'

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    danger_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    warning_font = Font(bold=True, color='C0392B')

    # Encabezados
    headers = [
        'ID', 'Nombre', 'Categoría', 'Proveedor',
        'Stock Actual', 'Stock Mínimo', 'Stock Requerido', 'Faltante'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos
    for row, p in enumerate(productos, 2):
        stock_requerido = p.stock_requerido or 0
        faltante = max(stock_requerido - p.cantidad_stock, 0)

        values = [
            p.id, p.nombre, p.categoria, p.proveedor or '-',
            p.cantidad_stock, p.stock_minimo, stock_requerido, faltante
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            # Resaltar stock actual en rojo
            if col == 5:
                cell.fill = danger_fill
                cell.font = warning_font

    # Auto-ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha = datetime.now().strftime('%Y-%m-%d')
    filename = f'alertas_stock_bajo_{fecha}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

