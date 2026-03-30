from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, send_file
from app.controllers.producto import ProductoController
from app.forms import ProductoForm
from app.services.alertas import verificar_stock_y_notificar, generar_informe_general
from app import db
from config import Config

producto_bp = Blueprint('producto', __name__, template_folder='../templates/producto')

def get_tasa_cambio():
    return Config.TASA_CAMBIO_USD_BS

@producto_bp.route('/')
def listar():
    from app.models import Producto
    
    # Obtener parámetros de búsqueda
    busqueda = request.args.get('busqueda', '').strip()
    categoria = request.args.get('categoria', '').strip()
    marca = request.args.get('marca', '').strip()
    precio_min = request.args.get('precio_min', type=float)
    precio_max = request.args.get('precio_max', type=float)
    precio_min_usd = request.args.get('precio_min_usd', type=float)
    precio_max_usd = request.args.get('precio_max_usd', type=float)
    proveedor = request.args.get('proveedor', '').strip()
    stock_bajo = request.args.get('stock_bajo', '').strip()
    stock_minimo = request.args.get('stock_minimo', type=int)
    stock_maximo = request.args.get('stock_maximo', type=int)
    
    # Verificar si hay filtros activos
    if busqueda or categoria or marca or precio_min or precio_max or precio_min_usd or precio_max_usd or proveedor or stock_bajo or stock_minimo or stock_maximo:
        productos = ProductoController.obtener_con_filtros(
            busqueda=busqueda if busqueda else None,
            categoria=categoria if categoria else None,
            marca=marca if marca else None,
            precio_min=precio_min,
            precio_max=precio_max,
            precio_min_usd=precio_min_usd,
            precio_max_usd=precio_max_usd,
            proveedor=proveedor if proveedor else None,
            stock_bajo=stock_bajo if stock_bajo else None,
            stock_minimo=stock_minimo,
            stock_maximo=stock_maximo
        )
    else:
        productos = ProductoController.obtener_todos()
    
    # Obtener categorías únicas para el filtro
    categorias = db.session.query(Producto.categoria).distinct().all()
    categorias = [c[0] for c in categorias]
    
    # Obtener marcas únicas para el filtro dinámico
    marcas = db.session.query(Producto.marca).filter(
        Producto.marca.isnot(None), Producto.marca != ''
    ).distinct().order_by(Producto.marca).all()
    marcas = [m[0] for m in marcas]
    
    return render_template('listar.html', productos=productos, categorias=categorias, marcas=marcas)

@producto_bp.route('/crear', methods=['GET', 'POST'])
def crear():
    form = ProductoForm()
    if form.validate_on_submit():
        data = {
            'nombre': form.nombre.data,
            'marca': form.marca.data,
            'descripcion': form.descripcion.data,
            'categoria': form.categoria.data,
            'cantidad_stock': form.cantidad_stock.data,
            'stock_minimo': form.stock_minimo.data,
            'proveedor': form.proveedor.data,
            'precio_mayor_bs': form.precio_mayor_bs.data or 0,
            'precio_mayor_usd': form.precio_mayor_usd.data or 0,
            'precio_detal_bs': form.precio_detal_bs.data or 0,
            'precio_detal_usd': form.precio_detal_usd.data or 0,
            'precio_tecnico_bs': form.precio_tecnico_bs.data or 0,
            'precio_tecnico_usd': form.precio_tecnico_usd.data or 0
        }
        modelos_ids = form.modelos_compatibles.data if form.categoria.data == 'pantalla' else []
        ProductoController.crear(data, modelos_ids)
        flash('Producto creado exitosamente', 'success')
        return redirect(url_for('producto.listar'))
    tasa = get_tasa_cambio()
    return render_template('crear.html', form=form, tasa_cambio=tasa)

@producto_bp.route('/editar/<int:id>', methods=['GET', 'POST'])
def editar(id):
    producto = ProductoController.obtener_por_id(id)
    if not producto:
        flash('Producto no encontrado', 'danger')
        return redirect(url_for('producto.listar'))

    form = ProductoForm(obj=producto)
    if request.method == 'GET':
        # Precargar modelos compatibles
        modelos_ids = [m.id for m in producto.modelos_compatibles]
        form.modelos_compatibles.data = modelos_ids
        print("=== CARGA INICIAL DEL FORMULARIO ===")
        print("Producto ID:", id)
        print("Modelos compatibles precargados (IDs):", modelos_ids)

    if form.validate_on_submit():
        print("=== FORMULARIO VÁLIDO ===")
        print("Categoría:", form.categoria.data)
        print("Modelos compatibles (enviados):", form.modelos_compatibles.data)
        data = {
            'nombre': form.nombre.data,
            'marca': form.marca.data,
            'descripcion': form.descripcion.data,
            'categoria': form.categoria.data,
            'cantidad_stock': form.cantidad_stock.data,
            'stock_minimo': form.stock_minimo.data,
            'proveedor': form.proveedor.data,
            'precio_mayor_bs': form.precio_mayor_bs.data or 0,
            'precio_mayor_usd': form.precio_mayor_usd.data or 0,
            'precio_detal_bs': form.precio_detal_bs.data or 0,
            'precio_detal_usd': form.precio_detal_usd.data or 0,
            'precio_tecnico_bs': form.precio_tecnico_bs.data or 0,
            'precio_tecnico_usd': form.precio_tecnico_usd.data or 0
        }
        modelos_ids = form.modelos_compatibles.data if form.categoria.data == 'pantalla' else []
        ProductoController.actualizar(id, data, modelos_ids)
        flash('Producto actualizado correctamente', 'success')
        return redirect(url_for('producto.listar'))
    else:
        if form.errors:
            print("=== ERRORES EN FORMULARIO ===")
            print(form.errors)

    tasa = get_tasa_cambio()
    return render_template('producto/editar.html', form=form, producto=producto, tasa_cambio=tasa)

@producto_bp.route('/eliminar/<int:id>', methods=['POST'])
def eliminar(id):
    if ProductoController.eliminar(id):
        flash('Producto eliminado', 'success')
    else:
        flash('Error al eliminar el producto', 'danger')
    return redirect(url_for('producto.listar'))

@producto_bp.route('/enviar-alerta-manual')
def enviar_alerta_manual():
    """Envía manualmente una alerta de stock bajo por email."""
    try:
        verificar_stock_y_notificar(current_app._get_current_object())
        flash('Alerta de stock bajo enviada exitosamente', 'success')
    except Exception as e:
        flash(f'Error al enviar la alerta: {str(e)}', 'danger')
        current_app.logger.error(f"Error en envío manual de alerta: {str(e)}")
    return redirect(url_for('producto.listar'))

@producto_bp.route('/enviar-informe-manual')
def enviar_informe_manual():
    """Envía manualmente el informe general de inventario por email."""
    try:
        generar_informe_general(current_app._get_current_object())
        flash('Informe general enviado exitosamente', 'success')
    except Exception as e:
        flash(f'Error al enviar el informe: {str(e)}', 'danger')
        current_app.logger.error(f"Error en envío manual de informe: {str(e)}")
    return redirect(url_for('producto.listar'))

@producto_bp.route('/actualizar-stock/<int:id>', methods=['POST'])
def actualizar_stock(id):
    nuevo_stock = request.form.get('nuevo_stock', type=int)
    if nuevo_stock is not None and nuevo_stock >= 0:
        if ProductoController.actualizar_stock_rapido(id, nuevo_stock):
            flash('Stock actualizado rápidamente', 'success')
        else:
            flash('Error al actualizar el stock', 'danger')
    else:
        flash('Cantidad de stock inválida', 'danger')
        
    return redirect(request.referrer or url_for('producto.listar'))

@producto_bp.route('/exportar-excel')
def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from datetime import datetime
    
    productos = ProductoController.obtener_todos()
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario de Productos'
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'ID', 'Nombre', 'Marca', 'Categoría', 'Descripción',
        'Stock', 'Stock Mínimo', 'Proveedor',
        'Precio Mayor (USD)', 'Precio Mayor (Bs)',
        'Precio Detal (USD)', 'Precio Detal (Bs)',
        'Precio Técnico (USD)', 'Precio Técnico (Bs)',
        'Modelos Compatibles'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos
    for row, p in enumerate(productos, 2):
        modelos = ', '.join([m.nombre for m in p.modelos_compatibles]) if p.modelos_compatibles else '-'
        
        values = [
            p.id, p.nombre, p.marca or '-', p.categoria, p.descripcion or '-',
            p.cantidad_stock, p.stock_minimo, p.proveedor or '-',
            float(p.precio_mayor_usd or 0), float(p.precio_mayor_bs or 0),
            float(p.precio_detal_usd or 0), float(p.precio_detal_bs or 0),
            float(p.precio_tecnico_usd or 0), float(p.precio_tecnico_bs or 0),
            modelos
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if isinstance(value, float):
                cell.number_format = '#,##0.00'
    
    # Auto-ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    fecha = datetime.now().strftime('%Y-%m-%d')
    filename = f'inventario_productos_{fecha}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )